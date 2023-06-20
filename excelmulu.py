#创建目录

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


def generate_table_of_contents(excel_file, output_file):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(excel_file)

    # 创建目录页
    toc_sheet = workbook.create_sheet(title="目录")

    # 遍历工作簿
    for sheet_name in workbook.sheetnames:
        if sheet_name == "目录":
            continue  # 跳过目录页

        sheet = workbook[sheet_name]

        # 获取表名和表注释
        table_name = sheet.cell(row=2, column=1).value
        table_comment = sheet.cell(row=2, column=2).value

        # 在目录页中创建超链接
        cell = toc_sheet.cell(row=toc_sheet.max_row + 1, column=1)
        cell.value = table_name

        # 创建超链接
        # hyperlink = Hyperlink(ref=f"'{sheet_name}'!A1", target=sheet_name + '!A1')
        hyperlink = Hyperlink(ref='', target='',location=f"'{sheet_name}'!A1")

        cell.hyperlink = hyperlink

        # 写入表注释
        toc_sheet.cell(row=toc_sheet.max_row, column=2).value = table_comment

    # 调整列宽
    toc_sheet.column_dimensions[get_column_letter(1)].width = 30
    toc_sheet.column_dimensions[get_column_letter(2)].width = 40

    # 删除默认工作表
    if "Sheet" in workbook.sheetnames:
        default_sheet = workbook["Sheet"]
        workbook.remove(default_sheet)

    # 保存 Excel 文件
    workbook.save(output_file)


# 指定输入和输出的 Excel 文件名
input_file = "output.xlsx"
output_file = "output1.xlsx"

# 生成目录页并保存 Excel 文件
generate_table_of_contents(input_file, output_file)

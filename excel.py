# 把表合并起来

import os
import openpyxl

# 输入文件夹路径
folder_path = "D:/pyt/tabletxt"

# 创建目标Excel文件
merged_wb = openpyxl.Workbook()
merged_ws = merged_wb.active
merged_ws.title = "表结构"

# 获取文件夹中的所有Excel文件路径
excel_files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]

# 处理每个Excel文件
for excel_file in excel_files:
    # 构建Excel文件路径
    excel_path = os.path.join(folder_path, excel_file)

    # 加载Excel文件
    wb = openpyxl.load_workbook(excel_path)

    # 获取Excel文件的第一个工作表
    sheet = wb.active

    # 获取Excel文件的名称，并去掉前缀"cfps6_ckg."
    sheet_name = excel_file.replace("cfps6_ckg.", "")

    # 创建新的工作表，并复制数据
    new_ws = merged_wb.create_sheet(title=sheet_name)
    for row in sheet.iter_rows(values_only=True):
        new_ws.append(row)

    # 关闭Excel文件
    wb.close()

    print("合并工作表:", sheet_name)

# 保存合并后的Excel文件
merged_excel_path = os.path.join(folder_path, "表结构.xlsx")
merged_wb.save(merged_excel_path)
merged_wb.close()

print("合并完成，文件已保存为:", merged_excel_path)

# 把oracle转换为mysql，失败了

import os
import re
import openpyxl
from openpyxl.utils.exceptions import IllegalCharacterError


def sanitize_sheet_name(sheet_name):
    # 替换特殊字符为下划线
    invalid_chars = ['[', ']', '*', '/', '\\', '?', ':']
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '_')
    return sheet_name


def extract_table_structure(sql_code):
    # 提取表名
    table_name_match = re.search(r"create table (\w+\.\w+)", sql_code, re.IGNORECASE)
    table_name = table_name_match.group(1) if table_name_match else None

    # 提取表注释
    table_comment_match = re.search(r"comment on table (\w+\.\w+)\s+is '(.+)'", sql_code, re.IGNORECASE)
    table_comment = table_comment_match.group(2) if table_comment_match else None

    # 提取列名和列注释
    column_matches = re.findall(r"comment on column (\w+\.\w+)\.(\w+)\s+is '(.+)'", sql_code, re.IGNORECASE)
    columns = []
    for match in column_matches:
        column_name = match[1]
        column_comment = match[2]
        columns.append((column_name, column_comment))

    return table_name, table_comment, columns


def process_folder(folder_path, output_file):
    # 创建Excel工作簿
    workbook = openpyxl.Workbook()

    # 遍历文件夹中的txt文件
    file_list = [file for file in os.listdir(folder_path) if file.endswith('.txt')]

    for file in file_list:
        file_path = os.path.join(folder_path, file)
        sheet_name = os.path.splitext(file)[0]  # 使用文件名作为工作表名称
        sheet_name = sanitize_sheet_name(sheet_name)  # 清理工作表名称中的特殊字符

        # 读取txt文件内容
        with open(file_path, 'r') as f:
            sql_code = f.read()

        # 提取表结构信息
        table_name, table_comment, columns = extract_table_structure(sql_code)

        try:
            # 创建工作表
            worksheet = workbook.create_sheet(title=sheet_name)

            # 写入表头
            worksheet.cell(row=1, column=1, value='表名')
            worksheet.cell(row=1, column=2, value='表注释')
            worksheet.cell(row=1, column=3, value='列名')
            worksheet.cell(row=1, column=4, value='列注释')

            # 写入表结构信息
            worksheet.cell(row=2, column=1, value=table_name)
            worksheet.cell(row=2, column=2, value=table_comment)

            row = 3
            for column in columns:
                worksheet.cell(row=row, column=3, value=column[0])
                worksheet.cell(row=row, column=4, value=column[1])
                row += 1

        except IllegalCharacterError:
            print(f"Illegal sheet name: {sheet_name}. Skipped.")

    # 删除默认的Sheet工作表
    del workbook["Sheet"]

    # 保存Excel文件
    workbook.save(output_file)


# 指定文件夹路径和输出文件名
folder_path = 'D:/pyt/tabletxt'
output_file = 'output.xlsx'

# 处理文件夹中的txt文件并生成Excel
process_folder(folder_path, output_file)
